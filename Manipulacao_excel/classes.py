from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart
from openpyxl.drawing.image import Image

class LeitorAcoes:
    def __init__(self, arquivoPath: str = ""):
        self.arquivoPath = arquivoPath
        self.dados = []

    def ler_dados(self):
        with open(f"dados/{self.arquivoPath}.txt", "r") as f:
            conteudo = f.readlines()
            self.dados = [x.replace("\n","").split(";") for x in conteudo]

class GerenciadorPlanilha:
    def __init__(self):
        self.workbook = None
        self.planilha = None
    
    def addPlanilha(self, titulo: str = ""):
        # Evitar de ter uma planilha extra ao instanciar o atributo workbook com o objeto Workbook()
        if self.planilha is None: 
            self.workbook = Workbook()
            self.workbook.active.title = titulo
            self.planilha = self.workbook.active
            
            return self.planilha
            
        newPlanilha = self.workbook.create_sheet(title=titulo)
        self.workbook.active = newPlanilha
        self.planilha = newPlanilha
        
        return newPlanilha
    
    def addLinha(self, dados: list = []):
        if self.planilha is not None:
            self.planilha.append(dados)
    
    def updateCelula(self, celula: str, valor):
        if self.planilha is not None:
            self.planilha[celula] = valor
    
    def mergeCells(self, celula_inicio: str, celula_fim: str):
        if self.planilha is not None:
            self.planilha.merge_cells(f"{celula_inicio}:{celula_fim}")
    
    def setCellFont(self, celula: str, font: Font):
        if self.planilha is not None:
            self.planilha[celula].font = font
    
    def setCellFill(self, celula: str, fill: PatternFill):
        if self.planilha is not None:
            self.planilha[celula].fill = fill
    
    def setCellAlignment(self, celula: str, alignment: Alignment):
        if self.planilha is not None:
            self.planilha[celula].alignment = alignment
    
    def setCellValue(self, celula: str, valor):
        if self.planilha is not None:
            self.planilha[celula].value = valor
    
    def addChart2Planilha(self, posicao: str, cod, ref_x, ref_y):
        chart = LineChart()
        chart.title = f"Cotações de {cod}"
        chart.x_axis.title = "Data"
        chart.y_axis.title = "Valor (R$)"
        chart.height = 14.82
        chart.width = 33.87
        
        referencia_cot = ref_y
        referencia_dat = ref_x
        chart.add_data(referencia_cot)
        chart.set_categories(referencia_dat)
        
        # Por algum motivo, quando tentava atribuir valores pelas séries do gráfico, o excel sempre reclamava de algum
        # corrompimento, então optei por setar o estilo do gráfico que já vem com cores pré-definidas.
        chart.style = 20
        
        # cotacoes = chart.series[0]
        # bb_inferior = chart.series[1]
        # bb_superior = chart.series[2]
        
        # cotacoes.graphicalProperties.line.width = 0
        # cotacoes.graphicalProperties.line.solidFill = 'oa55ab'
        
        # bb_inferior.graphicalProperties.line.width = 0
        # bb_inferior.graphicalProperties.line.solidFill = 'a61508'
        
        # bb_superior.graphicalProperties.line.width = 0
        # bb_superior.graphicalProperties.line.solidFill = '12a154'
        
        self.planilha.add_chart(chart, posicao)
    
    def addImagem(self, posicao: str, imagemPath: str):
        if self.planilha is not None:
            img = Image(imagemPath)
            self.planilha.add_image(img, posicao)
    
    def saveWorkbook(self, filename: str):
        self.workbook.save(filename)
        print("Planilha criada com sucesso: " + filename)