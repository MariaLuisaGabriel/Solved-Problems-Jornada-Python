#==============================================================
# JORNADA PYTHON (Projeto 2)
# Manipulação de arquivos Excel
#
# Uso do módulo requests para requisições HTTP e extração dos dados da tabela
# do site Fundamentus (https://fundamentus.com.br/fii_resultado.php), e com o módulo 
# BeautifulSoup faço o parsing do HTML.
#
# Após a extração, aplico uma estratégia de filtragem baseada em critérios financeiros
# e exponho os dados filtrados em uma tabela formatada com o módulo tabulate.
#==============================================================

import os, time
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from classes import LeitorAcoes, GerenciadorPlanilha

def clear_console():
    """Clears the console screen."""
    # For Windows
    if os.name == 'nt':
        _ = os.system('cls')
    # For macOS and Linux
    else:
        _ = os.system('clear')

arquivos = [x.split(".txt")[0] for x in os.listdir('dados')]
lim = len(arquivos)//11

while 1:
    print("Códigos de ações disponíveis para consulta:")
    for i in range(lim):
        print(f"- {arquivos[i]};")

    cod = input("Digite o código da ação que quer consultar ( \"m\" para mais opções de ação, \"x\" para sair )\n-> ")
    
    if cod == "x" or cod == "X":
        print("Saindo...")
        break
    
    if cod == "m" or cod == "M":
        lim *= 2
        if lim > len(arquivos):
            lim = len(arquivos)
    else:
        try:
            leitorAcoes = LeitorAcoes(cod)
            
            leitorAcoes.ler_dados()
            
            gerenciadorPlanilha = GerenciadorPlanilha()
            planilha_dados = gerenciadorPlanilha.addPlanilha(f"Dados {cod}")
            
            i = 2
            gerenciadorPlanilha.addLinha(["DATA", "COTAÇÃO", "BANDA INFERIOR", "BANDA SUPERIOR"])
            for linha in leitorAcoes.dados:
                
                ymd = linha[0].split(" ")[0].split("-")
                data = date(year = int(ymd[0]), month = int(ymd[1]), day = int(ymd[2]))
                
                cotacao = float(linha[1])
                
                gerenciadorPlanilha.updateCelula(f'A{i}', data)
                gerenciadorPlanilha.updateCelula(f'B{i}', cotacao)
                
                gerenciadorPlanilha.updateCelula(f'C{i}', cotacao)
                gerenciadorPlanilha.updateCelula(f'D{i}', cotacao)
                
                i+=1
            
            print(len(leitorAcoes.dados))
            print(i)
            
            for j in range(2, i):
                if j + 19 < i:
                    # banda inferior = média móvel 20 dias - desvio padrão 20 dias * 2
                    gerenciadorPlanilha.updateCelula(f'C{j}', f'=AVERAGE(B{j}:B{j+19}) - 2*STDEV(B{j}:B{j+19})')
                    
                    # banda superior = média móvel 20 dias + desvio padrão 20 dias * 2
                    gerenciadorPlanilha.updateCelula(f'D{j}', f'=AVERAGE(B{j}:B{j+19}) + 2*STDEV(B{j}:B{j+19})')
            
            gerenciadorPlanilha.addPlanilha("Gráfico")
            
            gerenciadorPlanilha.mergeCells("A1", "T2")
            
            gerenciadorPlanilha.setCellFont("A1", Font(bold=True, size=18, color="FFFFFF"))
            gerenciadorPlanilha.setCellFill("A1", PatternFill("solid", fgColor="07838f"))
            gerenciadorPlanilha.setCellAlignment("A1", Alignment(horizontal="center", vertical="center"))
            gerenciadorPlanilha.setCellValue("A1", f"Histórico de Cotações - {cod}")
            
            ref_y = Reference(planilha_dados, min_col=2, min_row=2, max_col=4, max_row=i)
            ref_x = Reference(planilha_dados, min_col=1, min_row=2, max_col=1, max_row=i)
            
            gerenciadorPlanilha.addChart2Planilha("A3", cod, ref_x, ref_y)
            
            gerenciadorPlanilha.mergeCells("I32", "L35")
            gerenciadorPlanilha.addImagem("I32", "./recursos/logo.png")
            
            gerenciadorPlanilha.saveWorkbook(f"./saida/planilha_{cod}.xlsx")
        except Exception as e:
            print(e)
    
        lim = len(arquivos)//6
        time.sleep(3)
        cod = input("\nConsultar outra ação? ( \"s\" Sim, \"n\" Não )\n-> ")
    
        if cod == "n" or cod == "N":
            print("Saindo...")
            break
        
    clear_console()